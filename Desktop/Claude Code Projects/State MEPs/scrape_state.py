import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import time
import re
import sys

def extract_profile_details(profile_url, headers):
    """Extract phone, mobile, email, and bio from individual profile page"""
    try:
        print(f"  Fetching profile: {profile_url}")
        response = requests.get(profile_url, headers=headers, timeout=30)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        phone = ""
        mobile = ""
        email = ""
        bio = ""

        # Look for contact information - try various patterns
        # Phone numbers are often in links with tel: or in text
        phone_links = soup.find_all('a', href=re.compile(r'tel:'))
        if phone_links:
            for link in phone_links:
                phone_text = link.get_text(strip=True)
                if 'mobile' in link.get('title', '').lower() or 'cell' in link.get('title', '').lower():
                    mobile = phone_text
                elif not phone:
                    phone = phone_text

        # Email - look for mailto links
        email_links = soup.find_all('a', href=re.compile(r'mailto:'))
        if email_links:
            email_href = email_links[0].get('href', '')
            email = email_href.replace('mailto:', '')

        # Bio - look for content area, article body, or bio section
        bio_areas = soup.find_all(['div', 'article', 'section'], class_=lambda x: x and ('content' in str(x).lower() or 'bio' in str(x).lower() or 'description' in str(x).lower() or 'text' in str(x).lower()))
        if bio_areas:
            for area in bio_areas:
                paragraphs = area.find_all('p')
                bio_text = ' '.join([p.get_text(strip=True) for p in paragraphs if p.get_text(strip=True)])
                if bio_text and len(bio_text) > 50:  # Make sure it's substantial
                    bio = bio_text
                    break

        # If phone/mobile/email not found in links, try to extract from bio text or page text
        page_text = soup.get_text()

        if not phone or not mobile:
            # Extract phone numbers from text
            phone_pattern = r'Phone:\s*\(?(\d{3})\)?[-.\s]?(\d{3})[-.\s]?(\d{4})'
            mobile_pattern = r'(?:Mobile|Cell):\s*\(?(\d{3})\)?[-.\s]?(\d{3})[-.\s]?(\d{4})'

            phone_match = re.search(phone_pattern, page_text)
            if phone_match and not phone:
                phone = f"({phone_match.group(1)}) {phone_match.group(2)}-{phone_match.group(3)}"

            mobile_match = re.search(mobile_pattern, page_text)
            if mobile_match and not mobile:
                mobile = f"({mobile_match.group(1)}) {mobile_match.group(2)}-{mobile_match.group(3)}"

        if not email:
            email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
            email_match = re.search(email_pattern, page_text)
            if email_match:
                email = email_match.group(0)

        return phone, mobile, email, bio
    except Exception as e:
        print(f"  Error fetching profile: {e}")
        return "", "", "", ""

def scrape_state_staff(state_name, staff_url):
    """Scrape staff information for a given state"""

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    try:
        print(f"\n{'='*60}")
        print(f"Processing: {state_name}")
        print(f"Staff Page: {staff_url}")
        print(f"{'='*60}\n")

        response = requests.get(staff_url, headers=headers, timeout=30)
        response.raise_for_status()

        soup = BeautifulSoup(response.content, 'html.parser')
        staff_data = []

        # Try multiple patterns to find staff members
        # Pattern 1: Team containers (like Arkansas)
        team_containers = soup.find_all('div', class_=lambda x: x and 'team' in str(x).lower())

        # Pattern 2: Staff/member cards
        if not team_containers:
            team_containers = soup.find_all(['div', 'article'], class_=lambda x: x and ('staff' in str(x).lower() or 'member' in str(x).lower() or 'person' in str(x).lower()))

        # Pattern 3: Look for names in headers with associated info
        if not team_containers:
            team_containers = soup.find_all(['div', 'section'], class_=lambda x: x and ('card' in str(x).lower() or 'profile' in str(x).lower()))

        print(f"Found {len(team_containers)} potential staff entries\n")

        if team_containers:
            for idx, container in enumerate(team_containers, 1):
                # Try to find name
                name = ""
                profile_url = ""
                title = ""

                # Look for name in various heading levels
                name_elem = container.find(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
                if name_elem:
                    name_link = name_elem.find('a')
                    if name_link:
                        name = name_link.get_text(strip=True)
                        profile_url = name_link.get('href', '')
                        if profile_url and not profile_url.startswith('http'):
                            # Make absolute URL
                            from urllib.parse import urljoin
                            profile_url = urljoin(staff_url, profile_url)
                    else:
                        name = name_elem.get_text(strip=True)

                # Look for title
                title_elem = container.find(['p', 'span', 'div'], class_=lambda x: x and ('title' in str(x).lower() or 'position' in str(x).lower() or 'role' in str(x).lower()))
                if title_elem:
                    title = title_elem.get_text(strip=True)
                elif name_elem:
                    # Title might be in next sibling
                    next_elem = name_elem.find_next(['p', 'span', 'div'])
                    if next_elem:
                        title = next_elem.get_text(strip=True)

                if name and len(name) < 100:  # Reasonable name length
                    print(f"[{idx}/{len(team_containers)}] Found: {name}")

                    phone = ""
                    mobile = ""
                    email = ""
                    bio = ""

                    # If there's a profile URL, fetch details
                    if profile_url:
                        phone, mobile, email, bio = extract_profile_details(profile_url, headers)
                        time.sleep(1)  # Be polite
                    else:
                        # Try to extract from current container
                        container_text = container.get_text()

                        # Phone
                        phone_match = re.search(r'(?:Phone|Tel|Office):\s*\(?(\d{3})\)?[-.\s]?(\d{3})[-.\s]?(\d{4})', container_text)
                        if phone_match:
                            phone = f"({phone_match.group(1)}) {phone_match.group(2)}-{phone_match.group(3)}"

                        # Mobile
                        mobile_match = re.search(r'(?:Mobile|Cell):\s*\(?(\d{3})\)?[-.\s]?(\d{3})[-.\s]?(\d{4})', container_text)
                        if mobile_match:
                            mobile = f"({mobile_match.group(1)}) {mobile_match.group(2)}-{mobile_match.group(3)}"

                        # Email
                        email_link = container.find('a', href=re.compile(r'mailto:'))
                        if email_link:
                            email = email_link.get('href', '').replace('mailto:', '')
                        else:
                            email_match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', container_text)
                            if email_match:
                                email = email_match.group(0)

                        # Bio
                        bio_paras = container.find_all('p')
                        bio = ' '.join([p.get_text(strip=True) for p in bio_paras if p.get_text(strip=True)])

                    staff_data.append({
                        'Name': name,
                        'Title': title,
                        'Phone': phone,
                        'Mobile': mobile,
                        'Email': email,
                        'Bio': bio
                    })
                    print(f"  Phone: {phone}, Mobile: {mobile}, Email: {email}")
                    print(f"  Complete\n")

        if staff_data:
            print(f"\nSuccessfully extracted {len(staff_data)} staff members")
            return staff_data
        else:
            print("\nNo staff data extracted. The page structure may need manual review.")
            return []

    except Exception as e:
        print(f"Error processing {state_name}: {e}")
        import traceback
        traceback.print_exc()
        return []

def update_excel_tab(state_abbrev, staff_data):
    """Update the Excel file with staff data for a specific state tab"""
    try:
        # Load the Excel workbook
        wb = load_workbook('state_meps.xlsx')

        # Access the state sheet
        if state_abbrev not in wb.sheetnames:
            print(f"Warning: Sheet {state_abbrev} not found in workbook")
            return False

        state_sheet = wb[state_abbrev]

        # Starting at Row 4
        start_row = 4

        for idx, staff in enumerate(staff_data):
            current_row = start_row + idx
            state_sheet[f'A{current_row}'] = staff['Name']
            state_sheet[f'B{current_row}'] = staff['Title']
            state_sheet[f'C{current_row}'] = staff['Phone'] if staff['Phone'] else ""
            state_sheet[f'D{current_row}'] = staff['Mobile'] if staff['Mobile'] else ""
            state_sheet[f'E{current_row}'] = staff['Email'] if staff['Email'] else ""
            state_sheet[f'F{current_row}'] = staff['Bio'] if staff['Bio'] else ""

        # Save the workbook
        wb.save('state_meps.xlsx')
        print(f"\nSuccessfully updated {state_abbrev} tab in state_meps.xlsx!")
        return True

    except Exception as e:
        print(f"Error updating Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python scrape_state.py <state_name> <state_abbrev> <staff_url>")
        sys.exit(1)

    state_name = sys.argv[1]
    state_abbrev = sys.argv[2]
    staff_url = sys.argv[3]

    # Scrape the staff data
    staff_data = scrape_state_staff(state_name, staff_url)

    # Update Excel if we got data
    if staff_data:
        # Save to temporary CSV for inspection
        df = pd.DataFrame(staff_data)
        csv_filename = f"{state_abbrev.lower()}_staff_temp.csv"
        df.to_csv(csv_filename, index=False)
        print(f"Saved to {csv_filename}")

        # Update Excel
        update_excel_tab(state_abbrev, staff_data)
    else:
        print(f"\nNo data to update for {state_name}")
