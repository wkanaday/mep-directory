from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import re

print("Setting up Chrome WebDriver for Delaware...")

chrome_options = Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('--no-sandbox')

driver = webdriver.Chrome(options=chrome_options)

# Try different pages
pages_to_check = [
    ('https://www.demep.org/about-us/', 'About Us'),
    ('https://www.demep.org/contact-us/', 'Contact Us'),
    ('https://www.demep.org/', 'Homepage')
]

staff_data = []

for url, page_name in pages_to_check:
    print(f"\nChecking {page_name} page: {url}")
    driver.get(url)
    time.sleep(5)

    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    # Look for staff information patterns
    # Check for email addresses with names
    text = soup.get_text()

    # Pattern 1: Name embedded in email (e.g., rachel.anderson@dtcc.edu)
    email_pattern = r'([a-z]+\.[a-z]+)@(?:demep\.org|dtcc\.edu)'
    emails_with_names = re.findall(email_pattern, text.lower())

    if emails_with_names:
        print(f"  Found {len(emails_with_names)} emails with name patterns")

    # Pattern 2: Look for structured staff listings
    staff_divs = soup.find_all(['div', 'section', 'article'], class_=lambda x: x and any(word in str(x).lower() for word in ['staff', 'team', 'member', 'person', 'employee', 'contact']))

    if staff_divs:
        print(f"  Found {len(staff_divs)} potential staff sections")

        for div in staff_divs[:10]:
            # Try to extract name, title, email
            div_text = div.get_text(strip=True)

            # Look for email in this section
            email_match = re.search(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.(?:org|edu|com))', div_text)
            if email_match:
                email = email_match.group(1)

                # Try to find name and title near the email
                # Look for headings
                name_elem = div.find(['h1', 'h2', 'h3', 'h4', 'strong', 'b'])
                name = name_elem.get_text(strip=True) if name_elem else ""

                # If no name yet, try to extract from email
                if not name and '@' in email:
                    name_part = email.split('@')[0]
                    # Convert firstname.lastname to Firstname Lastname
                    if '.' in name_part:
                        parts = name_part.split('.')
                        name = ' '.join(word.capitalize() for word in parts)

                # Look for title
                title = ""
                title_keywords = ['director', 'manager', 'president', 'ceo', 'coordinator', 'specialist', 'advisor']
                sentences = div_text.split('.')
                for sentence in sentences[:5]:
                    if any(keyword in sentence.lower() for keyword in title_keywords):
                        # Clean up the sentence to get just the title
                        clean_sentence = sentence.strip()
                        if len(clean_sentence) < 100:
                            title = clean_sentence
                            break

                if name:
                    staff_data.append({
                        'Name': name,
                        'Title': title,
                        'Phone': '',
                        'Mobile': '',
                        'Email': email,
                        'Bio': ''
                    })
                    print(f"    Found: {name} - {email}")

    # Pattern 3: Look for contact info in plain text
    # Split text into lines and look for name/title/email patterns
    lines = text.split('\n')
    for i, line in enumerate(lines):
        # Look for email addresses
        email_match = re.search(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.(?:org|edu|com))', line)
        if email_match:
            email = email_match.group(1)

            # Check surrounding lines for name and title
            context_lines = lines[max(0, i-3):min(len(lines), i+3)]
            context = ' '.join(context_lines)

            # Look for name pattern (capitalized words)
            name_match = re.search(r'([A-Z][a-z]+ [A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)', context)
            name = name_match.group(1) if name_match else ""

            # Look for title
            title = ""
            title_keywords = ['director', 'manager', 'president', 'ceo', 'coordinator', 'specialist', 'advisor', 'consultant']
            for keyword in title_keywords:
                if keyword in context.lower():
                    # Try to extract the full title
                    title_match = re.search(f'([^\\n]*{keyword}[^\\n]*)', context, re.IGNORECASE)
                    if title_match:
                        title = title_match.group(1).strip()[:100]
                        break

            if name and email:
                # Check if not already added
                if not any(s['Email'] == email for s in staff_data):
                    staff_data.append({
                        'Name': name,
                        'Title': title,
                        'Phone': '',
                        'Mobile': '',
                        'Email': email,
                        'Bio': ''
                    })
                    print(f"    Found in text: {name} - {email}")

driver.quit()

# Remove duplicates based on email
unique_staff = []
seen_emails = set()
for staff in staff_data:
    if staff['Email'] not in seen_emails:
        seen_emails.add(staff['Email'])
        unique_staff.append(staff)

staff_data = unique_staff

print(f"\n\nSuccessfully extracted {len(staff_data)} unique staff members")

if staff_data:
    for idx, staff in enumerate(staff_data, 1):
        print(f"{idx}. {staff['Name']} - {staff['Title']} - {staff['Email']}")

    # Save to CSV
    df = pd.DataFrame(staff_data)
    df.to_csv('de_staff_temp.csv', index=False)
    print("\nSaved to de_staff_temp.csv")

    # Update Excel
    print("\nUpdating DE tab in Excel...")
    try:
        wb = load_workbook('state_meps.xlsx')
        de_sheet = wb['DE']

        start_row = 4

        for idx, staff in enumerate(staff_data):
            current_row = start_row + idx
            de_sheet[f'A{current_row}'] = staff['Name']
            de_sheet[f'B{current_row}'] = staff['Title']
            de_sheet[f'C{current_row}'] = staff['Phone']
            de_sheet[f'D{current_row}'] = staff['Mobile']
            de_sheet[f'E{current_row}'] = staff['Email']
            de_sheet[f'F{current_row}'] = staff['Bio']

        wb.save('state_meps.xlsx')
        print(f"Successfully updated DE tab with {len(staff_data)} staff members!")
    except Exception as e:
        print(f"Error updating Excel: {e}")
        print("Please close the Excel file if it's open.")
else:
    print("\nNo staff data found. Delaware MEP may not have publicly listed staff information.")
    print("Creating empty entry to note this...")

    # Create a note entry
    try:
        wb = load_workbook('state_meps.xlsx')
        de_sheet = wb['DE']
        de_sheet['A4'] = 'No staff information publicly available'
        wb.save('state_meps.xlsx')
    except Exception as e:
        print(f"Error updating Excel: {e}")
